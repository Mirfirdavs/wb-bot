import os
import pandas as pd
from io import BytesIO
import asyncio
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, BufferedInputFile
from aiogram.client.default import DefaultBotProperties
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from dotenv import load_dotenv

# === Загрузка токена ===
load_dotenv()
TOKEN = os.getenv("BOT_TOKEN")
if not TOKEN:
    raise ValueError("BOT_TOKEN не найден в .env файле!")

# Инициализация бота с современными настройками
bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode="HTML"))
storage = MemoryStorage()
dp = Dispatcher(storage=storage)


# === FSM для выбора налога ===
class TaxState(StatesGroup):
    waiting_for_tax = State()


# === Глобальные переменные ===
user_data = {}  # Заменили глобальные переменные на словарь по user_id


# === Чтение Excel ===
def read_excel_bytes(file_bytes: BytesIO) -> pd.DataFrame:
    return pd.read_excel(file_bytes, engine="openpyxl")


# === Форматирование Excel ===
def apply_formats(output: BytesIO) -> BytesIO:
    wb = openpyxl.load_workbook(output)
    ws = wb.active

    financial_cols = [
        "Средняя цена розничная",
        "Среднее Вайлдберриз",
        "Выручка",
        "Сумма к перечислению",
        "Сумма услуг доставки",
        "Сумма штрафов",
        "Хранение",
        "Сумма WB Продвижение",
        "Налоги",
        "Себестоимость",
        "Итоговая себестоимость",
        "Чистая прибыль",
    ]
    percent_cols = ["Маржинальность", "Рентабельность"]

    headers = [cell.value for cell in ws[1]]
    financial_idx = [headers.index(c) + 1 for c in financial_cols if c in headers]
    percent_idx = [headers.index(c) + 1 for c in percent_cols if c in headers]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.value not in (None, ""):
                cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for idx in financial_idx:
            if idx - 1 < len(row):
                row[idx - 1].number_format = "#,##0.00 ₽"
        for idx in percent_idx:
            if idx - 1 < len(row):
                row[idx - 1].number_format = "0.0%"

    # Гистограмма Маржинальность
    if "Маржинальность" in headers:
        col_idx = headers.index("Маржинальность") + 1
        col_letter = get_column_letter(col_idx)
        rule = ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        )
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    # Гистограмма Рентабельность
    if "Рентабельность" in headers:
        col_idx = headers.index("Рентабельность") + 1
        col_letter = get_column_letter(col_idx)
        rule = ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        )
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# === Основной отчет ===
def process_main_report(file_bytes: BytesIO) -> pd.DataFrame:
    df = read_excel_bytes(file_bytes)
    df.columns = df.columns.str.strip()

    # Проверяем наличие необходимых колонок
    required_columns = ["Артикул поставщика", "Тип документа"]
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Отсутствует обязательная колонка: {col}")

    df = df[df["Артикул поставщика"].notna()]

    sales = df[df["Тип документа"].str.strip() == "Продажа"]
    sales_count = (
        sales.groupby("Артикул поставщика").size().reset_index(name="Количество продаж")
    )

    logistics = df[df["Обоснование для оплаты"].str.strip() == "Логистика"]
    logistics_count = (
        logistics.groupby("Артикул поставщика")
        .size()
        .reset_index(name="Количество доставок")
    )

    avg_price = (
        sales.groupby("Артикул поставщика")["Цена розничная"]
        .mean()
        .reset_index(name="Средняя цена розничная")
    )
    avg_wb = (
        sales.groupby("Артикул поставщика")["Вайлдберриз реализовал Товар (Пр)"]
        .mean()
        .reset_index(name="Среднее Вайлдберриз")
    )

    sum_payment = (
        sales.groupby("Артикул поставщика")[
            "К перечислению Продавцу за реализованный Товар"
        ]
        .sum()
        .reset_index(name="Сумма к перечислению")
    )
    sum_delivery = (
        logistics.groupby("Артикул поставщика")["Услуги по доставке товара покупателю"]
        .sum()
        .reset_index(name="Сумма услуг доставки")
    )
    sum_fines = (
        sales.groupby("Артикул поставщика")["Общая сумма штрафов"]
        .sum()
        .reset_index(name="Сумма штрафов")
    )
    storage = df.groupby("Артикул поставщика")["Хранение"].sum().reset_index()

    wb_promo = df[
        df["Виды логистики, штрафов и корректировок ВВ"].str.strip()
        == "Оказание услуг «WB Продвижение»"
    ]
    wb_promo_sum = (
        wb_promo.groupby("Артикул поставщика")["Вайлдберриз реализовал Товар (Пр)"]
        .sum()
        .reset_index(name="Сумма WB Продвижение")
    )

    result = (
        sales_count.merge(logistics_count, on="Артикул поставщика", how="left")
        .merge(avg_price, on="Артикул поставщика", how="left")
        .merge(avg_wb, on="Артикул поставщика", how="left")
        .merge(sum_payment, on="Артикул поставщика", how="left")
        .merge(sum_delivery, on="Артикул поставщика", how="left")
        .merge(sum_fines, on="Артикул поставщика", how="left")
        .merge(storage, on="Артикул поставщика", how="left")
        .merge(wb_promo_sum, on="Артикул поставщика", how="left")
    )

    revenue = (
        sales.groupby("Артикул поставщика")["Цена розничная"]
        .sum()
        .reset_index(name="Выручка")
    )
    result = result.merge(revenue, on="Артикул поставщика", how="left")

    result.fillna(0, inplace=True)
    cols = list(result.columns)
    if "Выручка" in cols and "Сумма к перечислению" in cols:
        cols.insert(cols.index("Сумма к перечислению"), cols.pop(cols.index("Выручка")))
        result = result[cols]

    return result


# === Обработка себестоимости ===
def process_cost(
    file_bytes: BytesIO, main_df: pd.DataFrame, tax_rate: float
) -> pd.DataFrame:
    df_cost = read_excel_bytes(file_bytes)
    df_cost.columns = df_cost.columns.str.strip()

    if (
        "Артикул поставщика" not in df_cost.columns
        or "Себестоимость" not in df_cost.columns
    ):
        raise ValueError(
            "В файле себестоимости отсутствуют необходимые колонки: 'Артикул поставщика' или 'Себестоимость'"
        )

    merged = pd.merge(
        main_df,
        df_cost[["Артикул поставщика", "Себестоимость"]],
        on="Артикул поставщика",
        how="left",
    )
    merged["Итоговая себестоимость"] = (
        merged["Себестоимость"] * merged["Количество продаж"]
    )
    merged["Налоги"] = merged["Сумма к перечислению"] * (tax_rate / 100)

    merged["Чистая прибыль"] = (
        merged["Сумма к перечислению"]
        - merged["Сумма услуг доставки"]
        - merged["Сумма штрафов"]
        - merged["Хранение"]
        - merged["Сумма WB Продвижение"]
        - merged["Налоги"]
        - merged["Итоговая себестоимость"]
    )

    merged["Маржинальность"] = merged["Чистая прибыль"] / merged["Выручка"].replace(
        0, 1
    )  # Избегаем деления на 0
    merged["Рентабельность"] = merged["Чистая прибыль"] / merged[
        "Итоговая себестоимость"
    ].replace(0, 1)

    merged.fillna(0, inplace=True)
    return merged


# === Генерация PDF с ТОП-4 и Прочее ===
def generate_short_pdf(df: pd.DataFrame) -> BytesIO:
    pdf_bytes = BytesIO()

    # --- Основная таблица ---
    pdf_df = df[
        [
            "Артикул поставщика",
            "Количество продаж",
            "Себестоимость",
            "Чистая прибыль",
            "Маржинальность",
            "Рентабельность",
        ]
    ].copy()
    pdf_df = pdf_df.sort_values(by="Чистая прибыль", ascending=False)
    pdf_df["Себестоимость"] = pdf_df["Себестоимость"].round(2)
    pdf_df["Чистая прибыль"] = pdf_df["Чистая прибыль"].round(1)
    pdf_df["Маржинальность"] = (pdf_df["Маржинальность"] * 100).round(1)
    pdf_df["Рентабельность"] = (pdf_df["Рентабельность"] * 100).round(1)

    totals = pd.DataFrame(
        {
            "Артикул поставщика": ["Итого"],
            "Количество продаж": [pdf_df["Количество продаж"].sum()],
            "Себестоимость": [pdf_df["Себестоимость"].sum()],
            "Чистая прибыль": [pdf_df["Чистая прибыль"].sum()],
            "Маржинальность": [pdf_df["Маржинальность"].mean()],
            "Рентабельность": [pdf_df["Рентабельность"].mean()],
        }
    )
    pdf_df = pd.concat([pdf_df, totals], ignore_index=True)

    with PdfPages(pdf_bytes) as pdf:
        # Основная таблица
        fig, ax = plt.subplots(figsize=(10, len(pdf_df) * 0.35 + 1))
        ax.axis("tight")
        ax.axis("off")
        display_df = pdf_df.copy()
        display_df["Себестоимость"] = display_df["Себестоимость"].apply(
            lambda x: f"{x:,.2f} ₽"
        )
        display_df["Чистая прибыль"] = display_df["Чистая прибыль"].apply(
            lambda x: f"{x:,.1f} ₽"
        )
        display_df["Маржинальность"] = display_df["Маржинальность"].apply(
            lambda x: f"{x:.1f}%"
        )
        display_df["Рентабельность"] = display_df["Рентабельность"].apply(
            lambda x: f"{x:.1f}%"
        )
        table = ax.table(
            cellText=display_df.values,
            colLabels=display_df.columns,
            loc="center",
            cellLoc="center",
        )
        table.auto_set_font_size(False)
        table.set_fontsize(8)
        table.auto_set_column_width(col=list(range(len(display_df.columns))))
        pdf.savefig(fig, bbox_inches="tight")
        plt.close()

        # --- ТОП источников дохода ---
        top_n = 4
        top_df = df.sort_values(by="Чистая прибыль", ascending=False).head(top_n)
        other_sum = (
            df.sort_values(by="Чистая прибыль", ascending=False)
            .iloc[top_n:]["Чистая прибыль"]
            .sum()
        )
        chart_df = top_df[["Артикул поставщика", "Чистая прибыль"]].copy()
        chart_df = pd.concat(
            [
                chart_df,
                pd.DataFrame(
                    [{"Артикул поставщика": "Прочее", "Чистая прибыль": other_sum}]
                ),
            ],
            ignore_index=True,
        )

        fig2, ax2 = plt.subplots(figsize=(8, 8))
        wedges, texts, autotexts = ax2.pie(
            chart_df["Чистая прибыль"],
            labels=chart_df["Артикул поставщика"],
            autopct="%1.1f%%",
            startangle=140,
            colors=plt.cm.tab20.colors,
        )
        ax2.set_title("ТОП источников дохода")
        pdf.savefig(fig2, bbox_inches="tight")
        plt.close()

    pdf_bytes.seek(0)
    return pdf_bytes


# === Кнопки налога ===
def tax_keyboard():
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="1%", callback_data="tax_1"),
                InlineKeyboardButton(text="6%", callback_data="tax_6"),
                InlineKeyboardButton(text="7%", callback_data="tax_7"),
            ],
            [
                InlineKeyboardButton(text="8%", callback_data="tax_8"),
                InlineKeyboardButton(text="Другое", callback_data="tax_other"),
            ],
        ]
    )
    return keyboard


# === Хэндлеры ===
@dp.message(Command("start"))
async def start(message: types.Message):
    # Инициализируем данные пользователя
    user_data[message.from_user.id] = {
        "main_df": None,
        "final_df": None,
        "tax_rate": 6,  # стандартная ставка по умолчанию
    }
    await message.answer(
        "👋 Привет! Выберите налоговую ставку:", reply_markup=tax_keyboard()
    )


@dp.callback_query(F.data.startswith("tax_"))
async def process_tax_choice(callback_query: types.CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    data = callback_query.data

    if user_id not in user_data:
        user_data[user_id] = {"main_df": None, "final_df": None, "tax_rate": 6}

    if data == "tax_other":
        await state.set_state(TaxState.waiting_for_tax)
        await callback_query.message.answer("Введите налоговую ставку (в % >0):")
    else:
        tax_rate = int(data.split("_")[1])
        user_data[user_id]["tax_rate"] = tax_rate
        await callback_query.message.answer(
            f"Вы выбрали налог: {tax_rate}%.\nОтправьте основной Excel-файл."
        )
        await state.clear()

    await callback_query.answer()


@dp.message(TaxState.waiting_for_tax)
async def process_custom_tax(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    if user_id not in user_data:
        user_data[user_id] = {"main_df": None, "final_df": None, "tax_rate": 6}

    try:
        tax = float(message.text.replace(",", "."))
        if tax <= 0:
            raise ValueError
        user_data[user_id]["tax_rate"] = tax
        await message.answer(
            f"Вы ввели налоговую ставку: {tax}%.\nТеперь отправьте основной Excel-файл."
        )
        await state.clear()
    except ValueError:
        await message.answer("❌ Некорректное значение. Введите число >0.")


@dp.message(F.document)
async def handle_file(message: types.Message):
    user_id = message.from_user.id

    if user_id not in user_data:
        await message.answer("Сначала нажмите /start")
        return

    try:
        doc = message.document
        file = await bot.get_file(doc.file_id)
        file_bytes = BytesIO()
        await bot.download_file(file.file_path, destination=file_bytes)
        file_bytes.seek(0)

        if "себестоимость" in doc.file_name.lower():
            if user_data[user_id]["main_df"] is None:
                await message.answer("⚠️ Сначала отправьте основной отчет!")
                return

            final_df = process_cost(
                file_bytes,
                user_data[user_id]["main_df"],
                user_data[user_id]["tax_rate"],
            )
            user_data[user_id]["final_df"] = final_df

            output = BytesIO()
            final_df.to_excel(output, index=False, engine="openpyxl")
            output = apply_formats(output)

            pdf_button = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="📄 Получить краткий PDF отчет",
                            callback_data="get_pdf",
                        )
                    ]
                ]
            )

            excel_file = BufferedInputFile(
                output.getvalue(), filename="Итоговый_отчет.xlsx"
            )
            await message.answer_document(excel_file)
            await message.answer("✅ Итоговый отчет готов!", reply_markup=pdf_button)
        else:
            main_df = process_main_report(file_bytes)
            user_data[user_id]["main_df"] = main_df
            await message.answer(
                "📊 Основной отчет обработан. Теперь отправьте файл себестоимости (.xlsx)."
            )

    except Exception as e:
        await message.answer(f"❌ Произошла ошибка: {str(e)}")


@dp.callback_query(F.data == "get_pdf")
async def send_short_pdf(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id

    if user_id not in user_data or user_data[user_id]["final_df"] is None:
        await callback_query.message.answer("⚠️ Сначала сформируйте итоговый отчет!")
        return

    try:
        pdf_file = generate_short_pdf(user_data[user_id]["final_df"])
        pdf_buffered = BufferedInputFile(
            pdf_file.getvalue(), filename="Краткий_отчет.pdf"
        )
        await callback_query.message.answer_document(pdf_buffered)
    except Exception as e:
        await callback_query.message.answer(f"❌ Ошибка при создании PDF: {str(e)}")


# === Запуск бота ===
async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
