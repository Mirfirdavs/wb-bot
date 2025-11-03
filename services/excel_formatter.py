import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    """Класс для профессионального форматирования Excel с распределением затрат"""

    @staticmethod
    def apply_business_formatting(df: pd.DataFrame) -> BytesIO:
        """Применение бизнес-форматирования к отчету"""
        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Аналитика")
            workbook = writer.book
            worksheet = writer.sheets["Аналитика"]

            # Стили форматирования
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            center_alignment = Alignment(horizontal="center", vertical="center")

            # Форматирование заголовков
            for cell in worksheet[1]:
                cell.alignment = center_alignment
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.fill = PatternFill(
                    start_color="2E86AB", end_color="2E86AB", fill_type="solid"
                )
                cell.font = Font(color="FFFFFF", bold=True)

            # Определение типов колонок
            financial_columns = [
                "Средняя цена розничная",
                "Среднее Вайлдберриз",
                "Выручка",
                "Сумма к перечислению",
                "Сумма услуг доставки",
                "Сумма штрафов",
                "Хранение",
                "Сумма WB Продвижение",
                "Платная приемка",
                "Налоги",
                "Себестоимость",
                "Итоговая себестоимость",
                "Чистая прибыль",
            ]

            percent_columns = ["Маржинальность", "Рентабельность"]

            # Форматирование данных
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = thin_border

                    # Форматирование финансовых колонок
                    column_name = (
                        df.columns[cell.column - 1]
                        if cell.column <= len(df.columns)
                        else None
                    )
                    if column_name in financial_columns:
                        cell.number_format = '#,##0.00" ₽"'
                    elif column_name in percent_columns:
                        cell.number_format = "0.00%"

            # Условное форматирование
            ExcelFormatter._apply_conditional_formatting(worksheet, df)

            # Автоматическая ширина колонок
            ExcelFormatter._auto_adjust_columns(worksheet)

        output.seek(0)
        return output

    @staticmethod
    def _apply_conditional_formatting(worksheet, df):
        """Применение условного форматирования"""
        if "Маржинальность" in df.columns:
            col_idx = df.columns.get_loc("Маржинальность") + 1
            col_letter = get_column_letter(col_idx)

            rule = ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8696B",
                mid_type="num",
                mid_value=0.1,
                mid_color="FFEB84",
                end_type="num",
                end_value=0.2,
                end_color="63BE7B",
            )
            worksheet.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{worksheet.max_row}", rule
            )

    @staticmethod
    def _auto_adjust_columns(worksheet):
        """Автоматическая настройка ширины колонок"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


class CostDistributor:
    """Класс для распределения общих затрат по артикулам пропорционально количеству продаж"""

    @staticmethod
    def distribute_costs(df: pd.DataFrame) -> pd.DataFrame:
        """
        Распределяет общие затраты (хранение, WB продвижение, платную приемку)
        по артикулам пропорционально количеству продаж.
        """
        if df.empty or "Количество продаж" not in df.columns:
            return df

        result_df = df.copy()
        total_sales = result_df["Количество продаж"].sum()

        # Если нет продаж — не перераспределяем
        if total_sales <= 0:
            return result_df

        # --- 1. ХРАНЕНИЕ ---
        if "Хранение" in result_df.columns:
            total_storage = result_df["Хранение"].sum()
            if total_storage > 0:
                result_df["Хранение"] = (
                    result_df["Количество продаж"] / total_sales
                ) * total_storage

        # --- 2. ПРОДВИЖЕНИЕ WB ---
        if "Сумма WB Продвижение" in result_df.columns:
            total_promo = result_df["Сумма WB Продвижение"].sum()
            if total_promo > 0:
                result_df["Сумма WB Продвижение"] = (
                    result_df["Количество продаж"] / total_sales
                ) * total_promo

        # --- 3. ПЛАТНАЯ ПРИЕМКА ---
        # Для нее в отчете может быть индивидуальное значение по каждому артикулу.
        # Поэтому перераспределяем ТОЛЬКО если она указана как одна общая сумма.
        if "Платная приемка" in result_df.columns:
            # Проверим, одинаковые ли значения у всех артикулов (т.е. сумма общая)
            unique_acceptances = result_df["Платная приемка"].nunique(dropna=True)
            if unique_acceptances == 1 and result_df["Платная приемка"].iloc[0] > 0:
                total_acceptance = result_df["Платная приемка"].sum()
                result_df["Платная приемка"] = (
                    result_df["Количество продаж"] / total_sales
                ) * total_acceptance
            # иначе оставляем как есть — уже распределена индивидуально

        return result_df


class AdvancedReportProcessor:
    """Расширенный обработчик отчетов с распределением затрат"""

    @staticmethod
    def process_main_report(file_bytes: BytesIO) -> pd.DataFrame:
        """Обработка основного отчета с учетом распределения затрат"""
        df = pd.read_excel(file_bytes, engine="openpyxl")
        df.columns = df.columns.str.strip()

        # Проверяем наличие необходимых колонок
        required_columns = ["Артикул поставщика", "Тип документа"]
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"Отсутствует обязательная колонка: {col}")

        df = df[df["Артикул поставщика"].notna()]

        # Продажи
        sales = df[df["Тип документа"].str.strip() == "Продажа"]
        sales_count = (
            sales.groupby("Артикул поставщика")
            .size()
            .reset_index(name="Количество продаж")
        )

        # Логистика
        logistics = df[df["Обоснование для оплаты"].str.strip() == "Логистика"]
        logistics_count = (
            logistics.groupby("Артикул поставщика")
            .size()
            .reset_index(name="Количество доставок")
        )

        # Средние значения
        avg_price = (
            sales.groupby("Артикул поставщика")["Цена розничная"]
            .mean()
            .reset_index(name="Средняя цена розничная")
        )
        avg_wb = (
            sales.groupby("Артикул поставщика")["Вайлдберриз реализовал Товар (Пр)"]
            .mean()
            .reset_index(name="Среднее Вайлдберриз")
        )

        # Суммы
        sum_payment = (
            sales.groupby("Артикул поставщика")[
                "К перечислению Продавцу за реализованный Товар"
            ]
            .sum()
            .reset_index(name="Сумма к перечислению")
        )
        sum_delivery = (
            logistics.groupby("Артикул поставщика")[
                "Услуги по доставке товара покупателю"
            ]
            .sum()
            .reset_index(name="Сумма услуг доставки")
        )
        sum_fines = (
            sales.groupby("Артикул поставщика")["Общая сумма штрафов"]
            .sum()
            .reset_index(name="Сумма штрафов")
        )

        # Хранение (пока собираем общую сумму, потом распределим)
        storage_total = df.groupby("Артикул поставщика")["Хранение"].sum().reset_index()

        # WB Продвижение (пока собираем общую сумму, потом распределим)
        wb_promo = df[
            df["Виды логистики, штрафов и корректировок ВВ"].str.strip()
            == "Оказание услуг «WB Продвижение»"
        ]
        wb_promo_sum = (
            wb_promo.groupby("Артикул поставщика")["Вайлдберриз реализовал Товар (Пр)"]
            .sum()
            .reset_index(name="Сумма WB Продвижение")
        )

        # Платная приемка (если есть в данных)
        paid_acceptance = df[
            df["Обоснование для оплаты"].str.strip() == "Платная приемка"
        ]
        paid_acceptance_sum = (
            paid_acceptance.groupby("Артикул поставщика")[
                "Услуги по доставке товара покупателю"
            ]
            .sum()
            .reset_index(name="Платная приемка")
        )

        # Собираем базовый результат
        result = (
            sales_count.merge(logistics_count, on="Артикул поставщика", how="left")
            .merge(avg_price, on="Артикул поставщика", how="left")
            .merge(avg_wb, on="Артикул поставщика", how="left")
            .merge(sum_payment, on="Артикул поставщика", how="left")
            .merge(sum_delivery, on="Артикул поставщика", how="left")
            .merge(sum_fines, on="Артикул поставщика", how="left")
            .merge(storage_total, on="Артикул поставщика", how="left")
            .merge(wb_promo_sum, on="Артикул поставщика", how="left")
        )

        # Добавляем платную приемку если есть данные
        if not paid_acceptance_sum.empty:
            result = result.merge(
                paid_acceptance_sum, on="Артикул поставщика", how="left"
            )

        # Выручка
        revenue = (
            sales.groupby("Артикул поставщика")["Цена розничная"]
            .sum()
            .reset_index(name="Выручка")
        )
        result = result.merge(revenue, on="Артикул поставщика", how="left")

        result.fillna(0, inplace=True)

        # Распределяем затраты пропорционально количеству продаж
        result = CostDistributor.distribute_costs(result)

        # Переупорядочиваем колонки
        cols = list(result.columns)
        if "Выручка" in cols and "Сумма к перечислению" in cols:
            cols.insert(
                cols.index("Сумма к перечислению"), cols.pop(cols.index("Выручка"))
            )
            result = result[cols]

        return result

    @staticmethod
    def process_cost(
        file_bytes: BytesIO, main_df: pd.DataFrame, tax_rate: float
    ) -> pd.DataFrame:
        """Обработка себестоимости с учетом распределенных затрат"""
        df_cost = pd.read_excel(file_bytes, engine="openpyxl")
        df_cost.columns = df_cost.columns.str.strip()

        if (
            "Артикул поставщика" not in df_cost.columns
            or "Себестоимость" not in df_cost.columns
        ):
            raise ValueError(
                "В файле себестоимости отсутствуют необходимые колонки: 'Артикул поставщика' или 'Себестоимость'"
            )

        merged = pd.merge(
            main_df,
            df_cost[["Артикул поставщика", "Себестоимость"]],
            on="Артикул поставщика",
            how="left",
        )
        merged["Итоговая себестоимость"] = (
            merged["Себестоимость"] * merged["Количество продаж"]
        )
        merged["Налоги"] = merged["Сумма к перечислению"] * (tax_rate / 100)

        # Расчет чистой прибыли с учетом ВСЕХ затрат
        cost_components = [
            "Сумма услуг доставки",
            "Сумма штрафов",
            "Хранение",
            "Сумма WB Продвижение",
            "Налоги",
            "Итоговая себестоимость",
        ]

        # Добавляем платную приемку если есть
        if "Платная приемка" in merged.columns:
            cost_components.insert(3, "Платная приемка")

        merged["Чистая прибыль"] = merged["Сумма к перечислению"]
        for cost_component in cost_components:
            if cost_component in merged.columns:
                merged["Чистая прибыль"] -= merged[cost_component]

        # Расчет маржинальности и рентабельности
        merged["Маржинальность"] = merged["Чистая прибыль"] / merged["Выручка"].replace(
            0, 1
        )
        merged["Рентабельность"] = merged["Чистая прибыль"] / merged[
            "Итоговая себестоимость"
        ].replace(0, 1)

        merged.fillna(0, inplace=True)
        return merged
